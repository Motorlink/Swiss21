#!/usr/bin/env python3
"""
Erstellt Excel-Datei mit Kundendaten und Aufträgen mit Netto- und Brutto-Beträgen
"""
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

print("=" * 80)
print("SCHRITT 1: Kundendaten aus CSV laden")
print("=" * 80)

kunden_daten = {}
with open('/home/ubuntu/upload/kundenliste.csv', 'r', encoding='ISO-8859-1') as f:
    reader = csv.DictReader(f, delimiter=';')
    
    for row in reader:
        kunde_nr = row.get('Nummer', '').strip()
        
        if not kunde_nr:
            continue
            
        kunden_daten[kunde_nr] = {
            'name': row.get('Name', '').strip(),
            'ansprechpartner': row.get('Name des Ansprechpartner', '').strip(),
            'email': row.get('E-Mail', '').strip(),
            'telefon': row.get('Telefon (gesch.)', '').strip(),
            'mobil': row.get('Mobil', '').strip(),
            'strasse': row.get('Straße', '').strip(),
            'plz': row.get('PLZ', '').strip(),
            'ort': row.get('Ort', '').strip(),
            'land': row.get('Land', '').strip(),
            'uid': row.get('USt-IdNr.', '').strip(),
            'kunde_seit': row.get('Kunde seit', '').strip(),
        }

print(f"✓ {len(kunden_daten)} Kunden geladen")

print("\n" + "=" * 80)
print("SCHRITT 2: Auftragsdaten aus JSON laden")
print("=" * 80)

with open('/tmp/abisco_auftraege_mit_listenpreis.json', 'r') as f:
    auftraege = json.load(f)

# Gruppiere Aufträge nach Kunde
auftraege_pro_kunde = defaultdict(list)
for auftrag in auftraege:
    kunde_nr = str(auftrag.get('kundennummer', ''))
    auftraege_pro_kunde[kunde_nr].append(auftrag)

print(f"✓ {len(auftraege)} Aufträge geladen")
print(f"✓ {len(auftraege_pro_kunde)} Kunden mit Aufträgen")

print("\n" + "=" * 80)
print("SCHRITT 3: Excel-Datei erstellen")
print("=" * 80)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # Entferne Standard-Sheet

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=11)
data_font = Font(size=10)
label_font = Font(bold=True, size=11)

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Für jeden Kunden mit Aufträgen ein Sheet erstellen
for kunde_nr in sorted(auftraege_pro_kunde.keys(), key=lambda x: int(x) if x.isdigit() else 0):
    kunde = kunden_daten.get(kunde_nr, {})
    kunde_name = kunde.get('name', f'Kunde {kunde_nr}')
    
    # Sheet-Name (max 31 Zeichen)
    sheet_name = f"{kunde_nr} {kunde_name}"[:31]
    ws = wb.create_sheet(title=sheet_name)
    
    print(f"Erstelle Sheet: {sheet_name}")
    
    # KUNDENDATEN
    ws['A1'] = 'KUNDENSTAMMDATEN'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    row = 2
    ws[f'A{row}'] = 'Kundennummer:'
    ws[f'B{row}'] = kunde_nr
    ws[f'A{row}'].font = label_font
    
    row += 1
    ws[f'A{row}'] = 'Firmenname:'
    ws[f'B{row}'] = kunde.get('name', '')
    ws[f'A{row}'].font = label_font
    
    row += 1
    ws[f'A{row}'] = 'Ansprechpartner:'
    ws[f'B{row}'] = kunde.get('ansprechpartner', '')
    ws[f'A{row}'].font = label_font
    
    row += 1
    ws[f'A{row}'] = 'Adresse:'
    adresse = f"{kunde.get('strasse', '')}, {kunde.get('plz', '')} {kunde.get('ort', '')}"
    if kunde.get('land'):
        adresse += f", {kunde.get('land')}"
    ws[f'B{row}'] = adresse
    ws[f'A{row}'].font = label_font
    
    row += 1
    ws[f'A{row}'] = 'E-Mail:'
    ws[f'B{row}'] = kunde.get('email', '')
    ws[f'A{row}'].font = label_font
    
    row += 1
    ws[f'A{row}'] = 'Telefon:'
    telefon = kunde.get('telefon', '')
    if kunde.get('mobil'):
        telefon += f" / {kunde.get('mobil')}"
    ws[f'B{row}'] = telefon
    ws[f'A{row}'].font = label_font
    
    row += 1
    ws[f'A{row}'] = 'UID-Nummer:'
    ws[f'B{row}'] = kunde.get('uid', '')
    ws[f'A{row}'].font = label_font
    
    # Leerzeile
    row += 2
    
    # AUFTRÄGE
    ws[f'A{row}'] = 'AUFTRÄGE'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    
    row += 1
    
    # Tabellen-Header
    headers = ['Auftrag', 'Datum', 'Einkaufspreis (CHF)', 'Betrag Netto (CHF)', 'Betrag Brutto (CHF)', 'Status', 'Status-Code']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Aufträge eintragen
    for auftrag in sorted(auftraege_pro_kunde[kunde_nr], key=lambda x: x.get('datum', '')):
        ws[f'A{row}'] = auftrag.get('auftragsnummer', '')
        ws[f'B{row}'] = auftrag.get('datum', '')
        ws[f'C{row}'] = float(auftrag.get('einkaufspreis_netto', 0))
        ws[f'D{row}'] = float(auftrag.get('betrag_netto', 0))
        ws[f'E{row}'] = float(auftrag.get('betrag_brutto', 0))
        
        # Status interpretieren (26 = abgeschlossen, 27 = offen)
        status_code = auftrag.get('status', 0)
        if status_code == 26:
            status_text = 'Abgeschlossen'
        elif status_code == 27:
            status_text = 'Offen'
        else:
            status_text = f'Status {status_code}'
        
        ws[f'F{row}'] = status_text
        ws[f'G{row}'] = status_code
        
        # Formatierung
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'E{row}'].number_format = '#,##0.00'
        
        # Borders
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Spaltenbreiten
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12

print("\n" + "=" * 80)
print("SCHRITT 4: Excel-Datei speichern")
print("=" * 80)

output_path = '/home/ubuntu/Swiss21/data/Kunden_Auftraege.xlsx'
wb.save(output_path)
print(f"✓ Datei gespeichert: {output_path}")

print("\n" + "=" * 80)
print("FERTIG!")
print("=" * 80)
print(f"✓ {len(wb.sheetnames)} Arbeitsblätter erstellt")
print(f"✓ {len(auftraege)} Aufträge aus Datenbank übernommen")
print("=" * 80)
