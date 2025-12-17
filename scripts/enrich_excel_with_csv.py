#!/usr/bin/env python3
"""
Ergänzt die Excel-Datei Kunden_Rechnungen.xlsx mit vollständigen Kundendaten aus der CSV
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# CSV-Datei einlesen
print("=" * 80)
print("SCHRITT 1: CSV-Daten einlesen")
print("=" * 80)

kunden_daten = {}
with open('/home/ubuntu/upload/kundenliste.csv', 'r', encoding='ISO-8859-1') as f:
    reader = csv.DictReader(f, delimiter=';')
    
    for row in reader:
        kunde_nr = row.get('Nummer', '').strip()
        
        # Wenn keine Kundennummer vorhanden, überspringe
        if not kunde_nr:
            continue
            
        kunden_daten[kunde_nr] = {
            'name': row.get('Name', '').strip(),
            'ansprechpartner': row.get('Name des Ansprechpartner', '').strip(),
            'email': row.get('E-Mail', '').strip(),
            'telefon': row.get('Telefon (gesch.)', '').strip(),
            'mobil': row.get('Mobil', '').strip(),
            'strasse': row.get('Straße', '').strip(),
            'plz': row.get('PLZ', '').strip(),
            'ort': row.get('Ort', '').strip(),
            'land': row.get('Land', '').strip(),
            'uid': row.get('USt-IdNr.', '').strip(),
            'typ': row.get('Typ', '').strip(),
            'kunde_seit': row.get('Kunde seit', '').strip(),
        }

print(f"✓ {len(kunden_daten)} Kunden aus CSV geladen")
for nr, data in list(kunden_daten.items())[:3]:
    print(f"  - {nr}: {data['name']}")

# Excel-Datei einlesen
print("\n" + "=" * 80)
print("SCHRITT 2: Excel-Datei öffnen und anreichern")
print("=" * 80)

wb = openpyxl.load_workbook('/home/ubuntu/Swiss21/data/Kunden_Rechnungen.xlsx')

# Durch alle Arbeitsblätter gehen
for sheet_name in wb.sheetnames:
    print(f"\nBearbeite Arbeitsblatt: {sheet_name}")
    ws = wb[sheet_name]
    
    # Extrahiere Kundennummer aus Sheet-Name
    # Format kann sein: "10000 Kunde 10000" oder "Kunde_10000"
    kunde_nr = None
    if ' Kunde ' in sheet_name:
        kunde_nr = sheet_name.split(' ')[0].strip()
    elif sheet_name.startswith('Kunde_'):
        kunde_nr = sheet_name.replace('Kunde_', '').strip()
    
    if kunde_nr and kunde_nr in kunden_daten:
        kunde = kunden_daten[kunde_nr]
        
        # Füge Kundendaten am Anfang des Arbeitsblatts ein
        # Verschiebe bestehende Daten nach unten
        ws.insert_rows(1, 10)
        
        # Header-Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)
        data_font = Font(size=11)
        label_font = Font(bold=True, size=11)
        
        # Kundendaten einfügen
        ws['A1'] = 'KUNDENSTAMMDATEN'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        ws[f'A{row}'] = 'Kundennummer:'
        ws[f'B{row}'] = kunde_nr
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        
        row += 1
        ws[f'A{row}'] = 'Firmenname:'
        ws[f'B{row}'] = kunde['name']
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        
        row += 1
        ws[f'A{row}'] = 'Ansprechpartner:'
        ws[f'B{row}'] = kunde['ansprechpartner']
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        
        row += 1
        ws[f'A{row}'] = 'Adresse:'
        adresse = f"{kunde['strasse']}, {kunde['plz']} {kunde['ort']}"
        if kunde['land']:
            adresse += f", {kunde['land']}"
        ws[f'B{row}'] = adresse
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        
        row += 1
        ws[f'A{row}'] = 'E-Mail:'
        ws[f'B{row}'] = kunde['email']
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        
        row += 1
        ws[f'A{row}'] = 'Telefon:'
        telefon = kunde['telefon']
        if kunde['mobil']:
            telefon += f" / {kunde['mobil']}"
        ws[f'B{row}'] = telefon
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        
        row += 1
        ws[f'A{row}'] = 'UID-Nummer:'
        ws[f'B{row}'] = kunde['uid']
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        
        row += 1
        ws[f'A{row}'] = 'Kunde seit:'
        ws[f'B{row}'] = kunde['kunde_seit']
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        
        # Spaltenbreiten anpassen
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        
        print(f"  ✓ Kundendaten für {kunde['name']} ({kunde_nr}) hinzugefügt")
    elif kunde_nr:
        print(f"  ⚠ Keine CSV-Daten für Kunde {kunde_nr} gefunden")
    else:
        print(f"  ⚠ Konnte Kundennummer nicht aus Sheet-Name extrahieren")

# Speichern
print("\n" + "=" * 80)
print("SCHRITT 3: Excel-Datei speichern")
print("=" * 80)

output_path = '/home/ubuntu/Swiss21/data/Kunden_Rechnungen_Vollstaendig.xlsx'
wb.save(output_path)
print(f"✓ Datei gespeichert: {output_path}")

print("\n" + "=" * 80)
print("FERTIG!")
print("=" * 80)
